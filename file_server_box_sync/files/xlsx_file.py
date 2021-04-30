from __future__ import annotations
import logging
from typing import AnyStr, Optional, List
import os

import openpyxl

from file_server_box_sync.files import sentinel_file


log = logging.getLogger(__name__)

"""
XLSX file
"""

class XLSXFile(sentinel_file.SentinelFile):

    __slots__ = [
        "attempted_upload_to_box_on",
        "box_file",
        "box_upload_folder_id",
        "box_upload_folder_path",
        "box_upload_user",
        "box_upload_user_email",
        "box_upload_user_id",
        "current_directory",
        "deleted",
        "deleted_on",
        "directory_entry_on",
        "file_directory_path",
        "file_name",
        "file_path",
        "headers",
        "id",
        "initial_monitor_on",
        "input_complete",
        "min_elapsed_for_box_upload_fail",
        "min_elapsed_for_delete",
        "min_elapsed_for_input_complete",
        "ready_for_delete",
        "ready_for_upload",
        "self_redis_key",
        "st_check_on",
        "st_creation_time",
        "st_last_accessed_time",
        "st_last_modified_time",
        "st_size",
        "st_size_diff_from_cache_on",
        "uploaded_to_box",
        "uploaded_to_box_on",
        "xlsx_workbook",
    ]

    def __init__(
        self,
        file_path: AnyStr,
        create_on_disk_on_init: bool = False,
        set_xlsx_workbook_on_init: bool = True,
        xlsx_workbook: Optional[openpyxl.reader.excel.ExcelReader] = None,
        **kwargs,
    ) -> XLSXFile:

        super().__init__(file_path=file_path, **kwargs)

        if create_on_disk_on_init:
            self.create_on_disk()

        self.xlsx_workbook = self.set_xlsx_workbook() if set_xlsx_workbook_on_init else xlsx_workbook

    def append_row(self, row_fields: List, worksheet: AnyStr = "Sheet") -> bool:
        appended = False
        try:
            self.xlsx_workbook[worksheet].append(row_fields)
            self.save_workbook()
            log.debug(f"{self} appended row {row_fields} to worksheet {worksheet}")
            appended = True
        except Exception as e:
            log.error(f"{self} failed to append row {row_fields} to sheet {worksheet} with {e}")

        return appended

    def create_on_disk(self, override_existing: bool = True) -> bool:
        created = False
        if override_existing or not os.path.exists(self.file_path):
            wb = openpyxl.Workbook()
            wb.save(self.file_path)
            created = True

        if created:
            log.debug(f"created file {str(self)} on disk")
        else:
            log.debug(f"did not create file {str(self)} on disk")

        return created

    def next_empty_row_index(self, worksheet_name: AnyStr = "Sheet") -> int:
        return self.xlsx_workbook[worksheet_name].max_row + 1

    def save_workbook(self, file_path: AnyStr = None) -> bool:
        saved = False
        try:
            if file_path:
                self.xlsx_workbook.save(filename=file_path)
            else:
                self.xlsx_workbook.save(filename=self.file_path)
            saved = True
        except Exception as e:
            error = e

        if saved:
            log.debug(f"{str(self)} save xlsx workbook")
        else:
            log.error(f"{str(self)} failed to save xlsx workbook with {str(error)}")

        return saved

    def set_xlsx_workbook(self, file_path: AnyStr = None) -> Optional[openpyxl.reader.excel.ExcelReader]:
        if not file_path:
            file_path = self.file_path

        xlsx_workbook = None
        if os.path.exists(file_path):
            self.xlsx_workbook = openpyxl.load_workbook(file_path)
            xlsx_workbook = self.xlsx_workbook

        return xlsx_workbook
